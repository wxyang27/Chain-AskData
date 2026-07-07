from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\r\n", "\n").replace("\r", "\n")
    lines = [" ".join(line.strip().split()) for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def split_list(value: str) -> list[str]:
    if not value:
        return []
    separators = ["、", "，", ",", "\n", "；", ";"]
    parts = [value]
    for separator in separators:
        parts = [piece for part in parts for piece in part.split(separator)]
    return [part.strip() for part in parts if part.strip()]


class IndicatorWorkbookLoader:
    """Load the 经管中心 indicator workbook into normalized asset dictionaries."""

    def load(self, workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
        workbook = load_workbook(workbook_path, data_only=True)
        return {
            "metrics": self._load_metrics(workbook_path, workbook),
            "user_profile_fields": self._load_user_profile_fields(
                workbook_path, workbook
            ),
            "dimensions": self._load_dimensions(workbook_path, workbook),
            "data_sources": self._load_data_sources(workbook_path, workbook),
            "dashboard_metrics": self._load_dashboard_metrics(
                workbook_path, workbook
            ),
        }

    def _rows(self, workbook_path: Path, workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
        sheet = workbook[sheet_name]
        headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        records: list[dict[str, Any]] = []

        for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            values = [clean_text(cell.value) for cell in row]
            if not any(values):
                continue
            record = dict(zip(headers, values, strict=False))
            record["_source_file"] = workbook_path.name
            record["_source_sheet"] = sheet_name
            record["_source_row"] = row_number
            records.append(record)

        return records

    def _load_metrics(self, workbook_path: Path, workbook: Any) -> list[dict[str, Any]]:
        metrics: list[dict[str, Any]] = []

        for row in self._rows(workbook_path, workbook, "原子指标字典"):
            metrics.append(
                {
                    "asset_id": f"metric:{row['指标ID']}",
                    "asset_type": "metric",
                    "metric_type": "atomic",
                    "metric_id": row["指标ID"],
                    "name": row["指标名称"],
                    "english_name": row.get("指标英文名", ""),
                    "business_domain": row.get("业务域", ""),
                    "definition": row.get("指标定义", ""),
                    "formula": row.get("计算逻辑", ""),
                    "data_type": row.get("数据类型", ""),
                    "unit": row.get("单位", ""),
                    "source_tables": split_list(row.get("数据源表", "")),
                    "sql": row.get("SQL代码", ""),
                    "derived_metric_count": row.get("衍生指标数", ""),
                    "dashboard_names": [],
                    "notes": row.get("备注", ""),
                    "source_file": row["_source_file"],
                    "source_sheet": row["_source_sheet"],
                    "source_row": row["_source_row"],
                }
            )

        for row in self._rows(workbook_path, workbook, "衍生指标字典"):
            metrics.append(
                {
                    "asset_id": f"metric:{row['指标ID']}",
                    "asset_type": "metric",
                    "metric_type": "derived",
                    "metric_id": row["指标ID"],
                    "name": row["指标名称"],
                    "english_name": "",
                    "business_domain": row.get("指标类型", ""),
                    "definition": row.get("指标定义", ""),
                    "formula": row.get("计算公式/逻辑", ""),
                    "data_type": "",
                    "unit": "",
                    "source_tables": split_list(row.get("数据源表", "")),
                    "sql": row.get("SQL代码", ""),
                    "atomic_metric_ids": split_list(row.get("对应原子指标", "")),
                    "dimension_slices": split_list(row.get("维度切片", "")),
                    "dashboard_names": split_list(row.get("所属看板", "")),
                    "notes": row.get("备注", ""),
                    "source_file": row["_source_file"],
                    "source_sheet": row["_source_sheet"],
                    "source_row": row["_source_row"],
                }
            )

        return metrics

    def _load_user_profile_fields(
        self, workbook_path: Path, workbook: Any
    ) -> list[dict[str, Any]]:
        fields = []
        for row in self._rows(workbook_path, workbook, "用户画像字段"):
            fields.append(
                {
                    "asset_id": f"user_profile_field:{row['字段ID']}",
                    "asset_type": "user_profile_field",
                    "field_id": row["字段ID"],
                    "name": row["字段名称"],
                    "business_area": row.get("业务板块", ""),
                    "category": row.get("子分类", ""),
                    "grain": row.get("数据粒度", ""),
                    "definition": row.get("字段定义", ""),
                    "data_type": row.get("数据类型", ""),
                    "source_tables": split_list(row.get("数据源表", "")),
                    "sql": row.get("SQL代码", ""),
                    "notes": row.get("备注", ""),
                    "source_file": row["_source_file"],
                    "source_sheet": row["_source_sheet"],
                    "source_row": row["_source_row"],
                }
            )
        return fields

    def _load_dimensions(self, workbook_path: Path, workbook: Any) -> list[dict[str, Any]]:
        dimensions = []
        for row in self._rows(workbook_path, workbook, "维度字典"):
            dimensions.append(
                {
                    "asset_id": f"dimension:{row['维度ID']}",
                    "asset_type": "dimension",
                    "dimension_id": row["维度ID"],
                    "name": row["维度名称"],
                    "english_name": row.get("维度英文名", ""),
                    "description": row.get("维度说明", ""),
                    "value_range": row.get("取值范围", ""),
                    "fields": split_list(row.get("对应字段", "")),
                    "source_tables": split_list(row.get("所属数据源", "")),
                    "affected_metrics": row.get("影响指标范围", ""),
                    "source_file": row["_source_file"],
                    "source_sheet": row["_source_sheet"],
                    "source_row": row["_source_row"],
                }
            )
        return dimensions

    def _load_data_sources(self, workbook_path: Path, workbook: Any) -> list[dict[str, Any]]:
        sources = []
        for row in self._rows(workbook_path, workbook, "数据源目录"):
            sources.append(
                {
                    "asset_id": f"data_source:{row['数据源ID']}",
                    "asset_type": "data_source",
                    "source_id": row["数据源ID"],
                    "table_name": row["表名"],
                    "alias": row.get("表简称", ""),
                    "database": row.get("库名", ""),
                    "business_description": row.get("业务说明", ""),
                    "main_fields": split_list(row.get("主要字段", "")),
                    "refresh_frequency": row.get("更新频率", ""),
                    "partition_field": row.get("分区字段", ""),
                    "estimated_metric_count": row.get("引用指标数(估)", ""),
                    "notes": row.get("备注", ""),
                    "source_file": row["_source_file"],
                    "source_sheet": row["_source_sheet"],
                    "source_row": row["_source_row"],
                }
            )
        return sources

    def _load_dashboard_metrics(
        self, workbook_path: Path, workbook: Any
    ) -> list[dict[str, Any]]:
        mappings = []
        for row in self._rows(workbook_path, workbook, "看板指标映射"):
            mappings.append(
                {
                    "asset_id": f"dashboard_metric:{row['序号']}",
                    "asset_type": "dashboard_metric",
                    "sequence": row["序号"],
                    "name": row["指标名称"],
                    "module": row.get("看板模块", ""),
                    "dashboard": row.get("所属看板", ""),
                    "dashboard_url": row.get("看板链接", ""),
                    "definition": row.get("指标定义", ""),
                    "owner": row.get("看板问题负责人", ""),
                    "in_metric_dictionary": row.get("是否在指标字典", ""),
                    "matched_metric_id": row.get("匹配指标ID", ""),
                    "source_file": row["_source_file"],
                    "source_sheet": row["_source_sheet"],
                    "source_row": row["_source_row"],
                }
            )
        return mappings

